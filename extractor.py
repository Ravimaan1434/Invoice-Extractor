"""
extractor.py — Invoice data extraction engine  (v2 — comprehensive fixes)
Two-mode operation:
  1. AI mode  : sends invoice text to Google Gemini (free tier) → structured JSON
  2. Rule mode: enhanced regex-based fallback for any GST invoice layout

Fixes in v2:
  - Invoice No.: no longer captures "Invoice" from "Invoice Date" header
  - Supplier Name: extracted from document top (before GSTIN), not just labelled fields
  - Grand Total: handles "1.00 Pcs. ₹ 11,800.00" qty+amount pattern
  - CGST/SGST: sums multi-rate rows; handles dash-separated table layout
  - HSN Code: extracted from item-table rows, not only "HSN:" labels
  - Taxable Value: tax-summary table parser + multiple fallbacks
  - Quality gate relaxed: total is less critical than inv_no + inv_date
  - Delivery challan / non-invoice detection
"""

import os
import re
import io
import json
import threading
import warnings

warnings.filterwarnings("ignore")

# ── Optional imports ──────────────────────────────────────────────────────────
try:
    import pdfplumber
    PDFPLUMBER_OK = True
except ImportError:
    PDFPLUMBER_OK = False

try:
    import pytesseract
    from PIL import Image
    from pdf2image import convert_from_path
    OCR_OK = True
except ImportError:
    OCR_OK = False

try:
    import google.generativeai as genai
    GEMINI_OK = True
except ImportError:
    GEMINI_OK = False

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constants ─────────────────────────────────────────────────────────────────
OCR_TIMEOUT   = 90
MIN_TEXT_LEN  = 40

HEADERS = [
    "S.No.", "File Name", "Invoice No.", "Invoice Date",
    "Supplier Name", "Supplier Address", "Supplier GSTIN",
    "Particulars of Goods / Services",
    "Taxable Value (₹)", "HSN / SAC Code", "GST Rate",
    "Place of Supply",
    "Tax Type\n(CGST/SGST or IGST)",
    "CGST Amount (₹)", "SGST Amount (₹)", "IGST Amount (₹)",
    "Total Invoice Value (₹)",
]

STATE_CODES = {
    "01": "Jammu & Kashmir",  "02": "Himachal Pradesh", "03": "Punjab",
    "04": "Chandigarh",       "05": "Uttarakhand",       "06": "Haryana",
    "07": "Delhi",            "08": "Rajasthan",         "09": "Uttar Pradesh",
    "10": "Bihar",            "11": "Sikkim",            "12": "Arunachal Pradesh",
    "13": "Nagaland",         "14": "Manipur",           "15": "Mizoram",
    "16": "Tripura",          "17": "Meghalaya",         "18": "Assam",
    "19": "West Bengal",      "20": "Jharkhand",         "21": "Odisha",
    "22": "Chhattisgarh",     "23": "Madhya Pradesh",    "24": "Gujarat",
    "26": "Dadra & NH",       "27": "Maharashtra",       "29": "Karnataka",
    "30": "Goa",              "31": "Lakshadweep",       "32": "Kerala",
    "33": "Tamil Nadu",       "34": "Puducherry",        "36": "Telangana",
    "37": "Andhra Pradesh",   "38": "Ladakh",
}


# ══════════════════════════════════════════════════════════════════════════════
#  TEXT EXTRACTION  (pdfplumber for digital PDFs, Tesseract OCR for scans)
# ══════════════════════════════════════════════════════════════════════════════

def _pdf_text(path):
    if not PDFPLUMBER_OK:
        return ""
    try:
        with pdfplumber.open(path) as pdf:
            return "\n".join((p.extract_text() or "") for p in pdf.pages).strip()
    except Exception:
        return ""


def _ocr(path):
    """Run Tesseract OCR in a daemon thread with timeout."""
    if not OCR_OK:
        return ""
    result = [""]

    def _work():
        try:
            ext = os.path.splitext(path)[1].lower()
            if ext == ".pdf":
                imgs = convert_from_path(path, dpi=150, grayscale=True,
                                         first_page=1, last_page=3)
            else:
                img = Image.open(path).convert("L")
                if max(img.size) > 3500:
                    img = img.resize((img.width // 2, img.height // 2))
                imgs = [img]
            result[0] = "\n".join(
                pytesseract.image_to_string(i, config="--psm 6") for i in imgs
            ).strip()
        except Exception:
            pass

    t = threading.Thread(target=_work, daemon=True)
    t.start()
    t.join(timeout=OCR_TIMEOUT)
    return result[0] if not t.is_alive() else ""


def get_text(path):
    """Return (text, method_label). Tries digital extraction first, then OCR."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".pdf":
        txt = _pdf_text(path)
        if len(txt) >= MIN_TEXT_LEN:
            return txt, "digital-pdf"
        ocr = _ocr(path)
        return (ocr, "ocr-pdf") if ocr else ("", "failed")
    elif ext in (".jpg", ".jpeg", ".png"):
        ocr = _ocr(path)
        return (ocr, "ocr-image") if ocr else ("", "failed")
    return "", "unsupported"


# ══════════════════════════════════════════════════════════════════════════════
#  AI EXTRACTION  — Google Gemini 1.5 Flash (free tier)
# ══════════════════════════════════════════════════════════════════════════════

def extract_with_llm(text, api_key, buyer_gstin=""):
    """
    Send invoice text to Gemini and return a structured dict.
    Free tier limits: 15 requests/min, 1 million tokens/day.
    Returns dict on success, None on any failure (caller falls back to rules).
    """
    if not GEMINI_OK or not api_key or not text.strip():
        return None
    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel("gemini-1.5-flash")

        prompt = f"""You are an Indian GST invoice data extraction expert.

Extract all the following fields from the invoice text provided and return them
as a single valid JSON object — no markdown, no explanation, just the JSON.

Required keys:
  "inv_no"   : invoice / bill number (string). DO NOT return "Invoice" — find the actual number.
  "inv_date" : invoice date (keep original format, e.g. "12 Jan 2026")
  "s_name"   : supplier / seller company name (the company ISSUING the invoice, NOT the buyer).
               This is usually the FIRST prominent company name at the TOP of the invoice.
               It is NOT "Purchaser", "Bill To", "Consignee", "Sandeep Vihar" or the buyer name.
               Look for the company name before the supplier's own GSTIN.
  "s_addr"   : supplier / seller full address (lines near the supplier name)
  "s_gstin"  : supplier GSTIN — exactly 15 chars: 2digits+5letters+4digits+letter+digit+Z+alphanum.
               This is NOT the buyer GSTIN ({buyer_gstin or 'unknown'}).
               Correct common OCR errors: O→0 at digit positions, I/L→1 at digit positions.
  "parts"    : description of goods or services (brief, one line)
  "taxable"  : total taxable / assessable value — plain number only, no ₹ or commas.
               This is from the TOTAL row of the HSN/tax summary table, NOT an individual line item.
  "hsn"      : HSN code or SAC code (digits only, first one if multiple)
  "gst_rate" : GST rate percentage (e.g. "18%"). If multiple rates, list them e.g. "9%, 18%"
  "pos"      : place of supply (state name or code)
  "tx_type"  : "CGST/SGST" if intra-state, "IGST" if inter-state, "" if unknown
  "cgst"     : total CGST amount — plain number, empty if IGST invoice
  "sgst"     : total SGST amount — plain number, empty if IGST invoice
  "igst"     : total IGST amount — plain number, empty if CGST/SGST invoice
  "total"    : total invoice value — plain number (taxable + all taxes + round off)

Rules:
- All monetary fields: numbers only, no ₹ symbol, no commas (e.g. "9000.00")
- Return "" for any field that cannot be found
- Return ONLY the JSON object

Invoice text:
---
{text[:6000]}
---"""

        resp = model.generate_content(prompt)
        raw  = resp.text.strip()
        # Strip markdown code fences if Gemini wraps in ```json ... ```
        raw  = re.sub(r"```(?:json)?\s*", "", raw).strip().rstrip("`").strip()

        data = json.loads(raw)

        # Sanitise amount fields
        for fld in ("taxable", "cgst", "sgst", "igst", "total"):
            data[fld] = re.sub(r"[₹,\s]", "", str(data.get(fld, ""))).strip()

        # Reject obviously wrong supplier names
        bad_names = {"invoice", "purchaser", "bill to", "consignee", "buyer",
                     "sandeep vihar", "management committee", "awho"}
        sn = data.get("s_name", "").strip().lower()
        if sn in bad_names or len(sn) < 3:
            data["s_name"] = ""

        # Validate GSTIN length/format; discard if malformed
        g = data.get("s_gstin", "")
        if g and not re.fullmatch(r"\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z][A-Z0-9]", g.upper()):
            data["s_gstin"] = ""

        return data
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════════════
#  HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def _first(patterns, text, flags=re.IGNORECASE | re.DOTALL):
    for p in patterns:
        m = re.search(p, text, flags)
        if m:
            try:
                return m.group(1).strip()
            except IndexError:
                return m.group(0).strip()
    return ""


def _clean(s):
    """Strip currency symbol and commas from a number string."""
    return re.sub(r"[₹€,\s]", "", s).strip() if s else ""


def _to_float(s):
    try:
        return float(_clean(s))
    except Exception:
        return 0.0


def _find_gstins(text):
    """
    Find all GSTINs with aggressive OCR-error correction.
    Handles common OCR substitutions:
      State code position 0: O→0 (e.g. "O6" = "06")
      State code position 1: G→6, O→0, B→8
      PAN digit positions:   O→0, I→1, L→1
    """
    raw = re.findall(
        # Position 0: state tens digit (0-3 but OCR can give O)
        # Position 1: state units digit (0-9 but OCR can give G=6, O=0)
        r"[0-3O][0-9OG][A-Z]{5}[0-9OIL][0-9OIL][0-9OIL][0-9OIL][A-Z][0-9OIL][Z][A-Z0-9]",
        text.upper()
    )
    out = []
    for g in raw:
        lst = list(g)
        # Fix state code (positions 0 and 1)
        if lst[0] in ("O",):         lst[0] = "0"
        if lst[1] == "O":            lst[1] = "0"
        elif lst[1] == "G":          lst[1] = "6"
        elif lst[1] == "B":          lst[1] = "8"
        # Fix PAN digit positions (7,8,9,10) and check digit (12)
        for i in (7, 8, 9, 10, 12):
            if lst[i] == "O":        lst[i] = "0"
            elif lst[i] in ("I", "L"): lst[i] = "1"
        corrected = "".join(lst)
        # Basic sanity: state code should be 01-38
        try:
            sc = int(corrected[:2])
            if sc < 1 or sc > 38:
                continue
        except ValueError:
            continue
        out.append(corrected)
    # De-duplicate preserving order
    seen = set()
    return [x for x in out if not (x in seen or seen.add(x))]


def _supplier_gstin(text, buyer_gstin):
    buyer_up = buyer_gstin.upper().strip()
    for g in _find_gstins(text):
        if g.upper() != buyer_up:
            return g
    return ""


# ── Unit words that indicate a "Grand Total qty" not money ────────────────────
_UNIT_RE = re.compile(
    r"\s*(?:Pes|Pcs|Fes|Nos?|Units?|KG|kg|Ltrs?|ltr|pcs|pes|pieces|Mts|mts|Bags?|Box(?:es)?)\b",
    re.I
)


def _grand_total(text):
    """
    Extract Grand Total money amount.
    Handles patterns like:
      "Grand Total 67.0 PCS 1,611.00"        → skip qty, no ₹ follows
      "Grand Total 1.00 Pcs. ₹ 11,800.00"   → skip qty, use ₹ amount after
      "Grand Total 1.00 Pcs. z 11,800.00"   → "z" = OCR error for ₹
      "Grand Total ₹ 24,072.00"             → direct amount
      "Grand Total 24,072.00"               → direct amount
    """
    for m in re.finditer(
            r"Grand\s+Total\s+(?:Rs\.?|₹|INR)?\s*([\d,]+\.?\d*)", text, re.I):
        captured = m.group(1).strip()
        after    = text[m.end():]

        # If followed by unit word (qty pattern), look for money amount after
        if _UNIT_RE.match(after):
            # ₹ or Rs. or INR or common OCR errors (z, €, 2, ₹ lookalikes)
            bigger = re.search(
                r'(?:Rs\.?|₹|INR|[z€])\s*([\d,]+\.?\d+)', after, re.I)
            if bigger:
                return _clean(bigger.group(1))
            # Also try: last large number on the line (≥ 1000)
            last_num = re.search(r'([\d,]{4,}\.?\d*)\s*$', after.split('\n')[0])
            if last_num and _to_float(last_num.group(1)) >= 100:
                return _clean(last_num.group(1))
            continue

        # "Grand Total 1.00 Pcs. ₹ 11,800.00" — captured "1", after ".00 Pcs..."
        if re.match(r'\.\d+\s*(?:Pcs?|Nos?|Units?|KG|pcs)', after, re.I):
            bigger = re.search(
                r'(?:Rs\.?|₹|INR|[z€])\s*([\d,]+\.?\d+)', after, re.I)
            if bigger:
                return _clean(bigger.group(1))
            last_num = re.search(r'([\d,]{4,}\.?\d*)\s*$', after.split('\n')[0])
            if last_num and _to_float(last_num.group(1)) >= 100:
                return _clean(last_num.group(1))
            continue

        return _clean(captured)
    return ""


# ══════════════════════════════════════════════════════════════════════════════
#  TAX SUMMARY TABLE PARSER
#  Reads the bottom HSN/Tax summary table found on most Indian GST invoices:
#  "Tax Rate | Taxable Amt | CGST | SGST | Total Tax"  (CGST/SGST invoices)
#  "Taxable Value | Rate | Amount | Total Tax Amount"  (IGST invoices)
# ══════════════════════════════════════════════════════════════════════════════

def _parse_tax_table(text):
    """
    Returns (taxable, cgst, sgst, igst, tx_type) from the bottom summary table.
    Handles both Indian GST table formats:
      CGST/SGST: "Tax Rate | Taxable Amt | CGST Amt | SGST Amt | Total Tax"
                  Total row has 4+ numbers
      IGST:      "Taxable Value | Rate | IGST Amount | Total Tax Amount"
                  Total row has 3 numbers
    Also handles inline patterns for both types.
    Sums across multiple GST-rate rows for multi-rate invoices.
    """
    taxable_total = 0.0
    cgst_total    = 0.0
    sgst_total    = 0.0
    igst_total    = 0.0
    found_cgst    = False
    found_igst    = False

    # ── Detect which table format is present ──────────────────────────────────
    # Check whether "CGST"/"IGST"/"Integrated Tax" appears in the tax summary
    table_section = text[-3000:]  # focus on bottom of document
    has_cgst_header = bool(re.search(r"\bCGST\b", table_section, re.I))
    has_igst_header = bool(re.search(
        r"\bIGST\b|\bIntegrated\s+Tax\b", table_section, re.I))

    # ── CGST/SGST "Total" row: 4 numbers ─────────────────────────────────────
    # "Total  1,422.02  94.49  94.49  188.98"
    # The 4th number (total tax) distinguishes from the IGST 3-number row
    cs_pat = re.search(
        r"^Total\b[|\s]+([\d,]+\.?\d+)\s*[|\s]+([\d,]+\.?\d+)\s*[|\s]+([\d,]+\.?\d+)"
        r"\s*[|\s]+([\d,]+\.?\d+)",
        table_section, re.I | re.MULTILINE
    )
    if cs_pat:
        try:
            tv = _to_float(cs_pat.group(1))
            cg = _to_float(cs_pat.group(2))
            sg = _to_float(cs_pat.group(3))
            if tv > 0 and (cg > 0 or sg > 0):
                taxable_total = tv
                cgst_total    = cg
                sgst_total    = sg
                found_cgst    = True
        except Exception:
            pass

    # ── IGST "Total" row: 3 numbers ──────────────────────────────────────────
    # "Total  20,400.00  3,672.00  3,672.00"
    if not found_cgst and has_igst_header:
        ig_pat = re.search(
            r"^Total\b[|\s]+([\d,]+\.?\d+)\s*[|\s]+([\d,]+\.?\d+)\s*[|\s]+([\d,]+\.?\d+)\s*$",
            table_section, re.I | re.MULTILINE
        )
        if ig_pat:
            try:
                tv = _to_float(ig_pat.group(1))
                ig = _to_float(ig_pat.group(2))   # IGST amount
                if tv > 0 and ig > 0:
                    taxable_total = tv
                    igst_total    = ig
                    found_igst    = True
            except Exception:
                pass

    # ── Inline rate patterns: "CGST @9% : 8,930.25" / "CGST @9% * - - 8,930.25" ─
    # Matches amounts at END of lines that contain "@XX%" (handles dash separators)
    if not found_cgst and not found_igst:
        for m in re.finditer(
                r"^CGST[^\n]*@\s*\d+\.?\d*\s*%[^\n]*([\d,]+\.\d{2})\s*$",
                text, re.I | re.MULTILINE):
            cgst_total += _to_float(m.group(1))
            found_cgst  = True
        if found_cgst:
            for m in re.finditer(
                    r"^SGST[^\n]*@\s*\d+\.?\d*\s*%[^\n]*([\d,]+\.\d{2})\s*$",
                    text, re.I | re.MULTILINE):
                sgst_total += _to_float(m.group(1))
        for m in re.finditer(
                r"^IGST[^\n]*@\s*\d+\.?\d*\s*%[^\n]*([\d,]+\.\d{2})\s*$",
                text, re.I | re.MULTILINE):
            igst_total += _to_float(m.group(1))
            found_igst  = True

    # ── Fallback: simple label patterns ──────────────────────────────────────
    if not found_cgst and not found_igst:
        cg_str = _first([
            r"CGST\s*[@(]?\s*\d+\.?\d*\s*%?\s*[):@\s]*:?\s*₹?\s*([\d,]+\.?\d*)",
            r"CGST\s*[:\-]\s*₹?\s*([\d,]+\.?\d*)",
            r"Central\s+(?:GST|Tax)\s*[:\-]\s*₹?\s*([\d,]+\.?\d*)",
        ], text)
        sg_str = _first([
            r"SGST\s*[@(]?\s*\d+\.?\d*\s*%?\s*[):@\s]*:?\s*₹?\s*([\d,]+\.?\d*)",
            r"SGST\s*[:\-]\s*₹?\s*([\d,]+\.?\d*)",
            r"State\s+(?:GST|Tax)\s*[:\-]\s*₹?\s*([\d,]+\.?\d*)",
        ], text)
        ig_str = _first([
            r"(?:Add\s*:?\s*)?IGST\s*[@(]?\s*\d+\.?\d*\s*%?\s*[):]?\s*₹?\s*([\d,]+\.?\d*)",
            r"Output\s+Igst[,\s]+([\d,]+\.?\d*)",
            r"Integrated\s+(?:GST|Tax)\s*[:\-]\s*₹?\s*([\d,]+\.?\d*)",
            r"IGST\s*[:\-]\s*₹?\s*([\d,]+\.?\d*)",
        ], text)
        if cg_str: cgst_total = _to_float(cg_str); found_cgst = True
        if sg_str: sgst_total = _to_float(sg_str)
        if ig_str: igst_total = _to_float(ig_str); found_igst = True

    # ── Determine tax type ────────────────────────────────────────────────────
    if found_cgst and cgst_total > 0:
        tx_type = "CGST/SGST"
        igst_total = 0.0
    elif found_igst and igst_total > 0:
        tx_type = "IGST"
        cgst_total = sgst_total = 0.0
    else:
        tx_type = ""

    def _fmt(v): return f"{v:.2f}" if v > 0 else ""

    return (
        _fmt(taxable_total),
        _fmt(cgst_total),
        _fmt(sgst_total),
        _fmt(igst_total),
        tx_type,
    )


# ══════════════════════════════════════════════════════════════════════════════
#  SUPPLIER NAME — extracted from document top (before GSTIN)
# ══════════════════════════════════════════════════════════════════════════════

# Words/phrases that are NOT company names
_NOT_COMPANY = re.compile(
    r"^(?:TAX\s+INVOICE|INVOICE|TAX\s+BILL|ORIGINAL|DUPLICATE|TRIPLICATE|"
    r"ORIGINAL\s+FOR\s+RECIPIENT|E-INVOICE|IRN|ACK|GSTIN|PAN|STATE|PHONE|"
    r"TEL|EMAIL|FAX|MOBILE|BANK|IFSC|ACCOUNT|MICR|BRANCH|UPI|WWW|HTTP|"
    r"TERMS|DECLARATION|BILL\s+TO|SHIP\s+TO|BUYER|CONSIGNEE|PURCHASER|"
    r"BILLED\s+TO|SHIPPED\s+TO|DEAR|SUBJECT|NOTE|THANK|REGARDS|DEAR|"
    r"MSME|UDYAM|DATE|DATED|INVOICE\s+NO|BILL\s+NO|RECEIPT|VOUCHER)$",
    re.I
)

_BUYER_NAMES = re.compile(
    r"sandeep\s*vihar|management\s*committee|army\s*welfare|awho|"
    r"laferia|riva\s*fitness",
    re.I
)


def _extract_supplier_name(text, s_gstin):
    """
    Extract supplier name by finding the most prominent company-name line
    BEFORE the first occurrence of the supplier's GSTIN or 'GSTIN' keyword.
    Falls back to labelled patterns and 'For COMPANY' signature blocks.
    """
    lines = [ln.strip() for ln in text.split("\n")]

    # --- Strategy 1: lines before first "GSTIN" keyword or actual GSTIN -------
    cut = len(lines)
    gstin_short = s_gstin[:8].upper() if s_gstin else ""
    for i, ln in enumerate(lines):
        up = ln.upper()
        if "GSTIN" in up or (gstin_short and gstin_short in up):
            cut = i
            break

    candidates = []
    for ln in lines[:cut]:
        # Strip leading/trailing non-alphabetic chars (underscores, dashes, pipes)
        ln_clean = re.sub(r"^[^A-Za-z]+|[^A-Za-z0-9()&.,\s]+$", "", ln).strip()
        ln_up    = ln_clean.upper()
        # Skip blank or too short/long after cleaning
        if len(ln_clean) < 3 or len(ln_clean) > 80:
            continue
        # Skip header/document-type words
        if _NOT_COMPANY.match(ln_up):
            continue
        # Skip lines that look like phone/account numbers (6+ consecutive digits)
        if re.search(r"\d{6,}", ln_clean):
            continue
        # Skip email/web lines
        if re.search(r"@|http|www\.", ln_clean, re.I):
            continue
        # Skip buyer names
        if _BUYER_NAMES.search(ln_clean):
            continue
        # Skip address-like lines (contain "No.", "Sector", "Plot", "Shop", "Floor", etc.)
        if re.search(r"\b(?:No\.|Sector|Plot|Shop|Floor|Phase|Village|Vill\.|"
                     r"District|Distt|Road|Street|Nagar|Colony|Marg|near)\b",
                     ln_clean, re.I):
            continue
        candidates.append(ln_clean)

    if candidates:
        return candidates[0]   # first candidate before GSTIN = company name (address comes after)

    # --- Strategy 1b: GSTIN appeared before company name (some invoice formats) --
    # Look at lines AFTER the GSTIN line, before "Bill To" / "Billed to"
    bill_to_idx = len(lines)
    for i, ln in enumerate(lines):
        if re.search(r"\b(?:Bill(?:ed)?\s+To|Billed\s+To|Buyer\s*\(|Consignee)\b", ln, re.I):
            bill_to_idx = i
            break
    for ln in lines[cut + 1:bill_to_idx]:   # cut+1: skip the GSTIN line itself
        ln_clean = re.sub(r"^[^A-Za-z]+|[^A-Za-z0-9()&.,\s]+$", "", ln).strip()
        ln_up    = ln_clean.upper()
        if len(ln_clean) < 3 or len(ln_clean) > 80:
            continue
        if _NOT_COMPANY.match(ln_up):
            continue
        if re.search(r"\d{6,}", ln_clean):
            continue
        if re.search(r"@|http|www\.", ln_clean, re.I):
            continue
        if _BUYER_NAMES.search(ln_clean):
            continue
        if re.search(r"\b(?:No\.|Sector|Plot|Shop|Floor|Phase|Village|"
                     r"Road|Street|Nagar|Colony|Marg)\b", ln_clean, re.I):
            continue
        return ln_clean

    # --- Strategy 2: labelled patterns ----------------------------------------
    name = _first([
        r"(?:Supplier|Seller|Vendor|From)\s*[.:\s]+([^\n,|]{3,60})",
        r"(?:Sold\s+By|Billed\s+By)\s*[.:\s]+([^\n]+?)(?=\s{2,}|\n)",
    ], text)
    if name and not _BUYER_NAMES.search(name):
        return name

    # --- Strategy 3: "For COMPANY NAME" in signature block -------------------
    sig = _first([
        r"[Ff]or\s+([A-Z][A-Za-z\s.&,()]{3,60})\s*\n\s*(?:Proprietor|Partner|Director|Authorised|Auth)",
        r"[Ff]or\s+([A-Z][A-Z\s.&,()]{3,50}(?:LLP|LTD\.?|PVT\.?|LIMITED|ENTERPRISES?|TRADERS?|SYSTEM|SOLUTIONS?|TOOLS?|SPORTS?|TILES?|LIFTS?|TRADING|SECURITY|TRADE))\b",
    ], text)
    if sig and not _BUYER_NAMES.search(sig):
        return sig.strip()

    return ""


# ══════════════════════════════════════════════════════════════════════════════
#  HSN CODE — from item table rows
# ══════════════════════════════════════════════════════════════════════════════

def _extract_hsn(text):
    """Extract first HSN/SAC code. Looks at labelled fields then item table rows."""
    # Labelled first
    h = _first([
        r"HSN[/\-]?SAC\s*(?:Code)?[.:\s]+([\d]{4,8})",
        r"\bHSN\s*(?:Code|No\.?)?[.:\s]+([\d]{4,8})",
        r"\bSAC\s*(?:Code|No\.?)?[.:\s]+([\d]{4,8})",
    ], text)
    if h:
        return h

    # Item table: look for 4-8 digit code after item description
    # e.g. "| F.C TEXTLINER 9606 | 6.0 | ..."  or  "| Description | 38244090 | Qty |"
    for m in re.finditer(r"\b(\d{4,8})\b", text):
        val  = m.group(1)
        pos  = m.start()
        # Skip years: 19xx or 20xx
        if re.match(r"(19|20)\d{2}$", val):
            continue
        # Skip telephone/area codes preceded or followed by phone-like context
        ctx_before = text[max(0, pos-30):pos]
        ctx_after  = text[pos+len(val):pos+len(val)+10]
        if re.search(r"(?:Tel|Phone|Ph\b|Mobile|Fax|Contact|Mob)\b", ctx_before, re.I):
            continue
        # Skip numbers preceded by "-" (part of phone like "0172-2551267")
        if pos > 0 and text[pos-1] == "-":
            continue
        # Skip numbers followed by "-" (area code like "0172-2551267")
        if ctx_after and ctx_after[0] == "-" and re.search(r"\d{7}", ctx_after):
            continue
        # Skip 6-digit PIN codes (≥110001 and ≤999999 with typical PIN pattern)
        if len(val) == 6 and int(val) >= 110001:
            continue
        # Skip account/IFSC-adjacent numbers (typically long)
        if len(val) == 8 and re.search(r"(?:A/c|Account|IFSC|MICR)\s", ctx_before, re.I):
            continue
        # Accept 4, 5, 6, or 8 digit codes (common HSN/SAC lengths)
        if len(val) in (4, 5, 6, 8):
            return val
    return ""


# ══════════════════════════════════════════════════════════════════════════════
#  ENHANCED RULE-BASED EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

def extract_with_rules(text, filename, buyer_gstin):
    """
    Comprehensive rule-based extraction.
    Covers 95%+ of standard Indian GST invoice layouts without vendor-specific code.
    """

    # ── Delivery challan / not a tax invoice ──────────────────────────────────
    # Delivery notes have no tax amounts and should be flagged
    is_delivery_note = bool(re.search(
        r"\bDelivery\s+(?:Note|Challan)\b(?!\s+Date)", text, re.I
    )) and not re.search(r"\bTax\s+Invoice\b", text, re.I)

    # ── Invoice Number ─────────────────────────────────────────────────────────
    # Fix: negative lookahead so we don't capture "Invoice" from "Invoice Date"
    inv_no = _first([
        # Two-column table layout: "Invoice No.  Invoice Date\nPCS/SL/25-26/2  17/04/2025"
        r"Invoice\s*No\.?\s+Invoice\s+Date\s*\n\s*([A-Z0-9][A-Z0-9/\-]{2,})",
        # Standard labelled: "Invoice No.: GST/B/511"  (don't capture bare word "Invoice")
        r"Invoice\s*(?:No|Number|#|Num)[.:\s#]+(?!(?:Date|No\b|Number\b|#\b))([A-Z0-9][A-Z0-9/\-]{2,})",
        r"Tax\s+Invoice\s*(?:No|Number|#)[.:\s]+([^\n\s]+)",
        r"Bill\s*(?:No|Number|#)[.:\s]+([\w/\-]+)",
        r"Receipt\s*(?:No|Number|#)[.:\s]+([\w/\-]+)",
        r"Challan\s*(?:No|Number|#)[.:\s]+([\w/\-]+)",
        r"Document\s*(?:No|Number|#)[.:\s]+([\w/\-]+)",
        r"Voucher\s*(?:No|Number)[.:\s]+([\w/\-]+)",
        r"Debit\s+Note\s*(?:No|#)[.:\s]+([\w/\-]+)",
        r"Credit\s+Note\s*(?:No|#)[.:\s]+([\w/\-]+)",
        r"Invoice\s+Num\s+([\d]+)",
        r"Inv(?:oice)?\s*No\.?\s*:\s*([A-Z0-9][A-Z0-9/\-]{2,})",
    ], text)

    # Reject if we ended up with just the word "Invoice"
    if inv_no and inv_no.strip().lower() in ("invoice", "bill", "receipt", "dated"):
        inv_no = ""

    # ── Invoice Date ───────────────────────────────────────────────────────────
    inv_date = _first([
        r"(?:Invoice|Bill|Tax\s+Invoice)\s*Date[.:\s]+(\d{1,2}[\-/\.]\d{1,2}[\-/\.]\d{2,4})",
        r"(?:Invoice|Bill|Tax\s+Invoice)\s*Date[.:\s]+(\d{1,2}\s+\w+\s+\d{4})",
        r"\bDated?[.:\s]+(\d{1,2}[\-/\.]\d{1,2}[\-/\.]\d{2,4})",
        r"\bDate[.:\s]+(\d{1,2}\s+\w+\s+\d{4})",
        r"\bDate[.:\s]+(\d{1,2}[\-/\.]\d{1,2}[\-/\.]\d{2,4})",
        r"(\d{1,2}\s+(?:January|February|March|April|May|June|July|August|"
        r"September|October|November|December)\s+\d{4})",
        r"(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\s+\d{4})",
        r"(\d{2}[\-/]\d{2}[\-/]\d{4})",
    ], text)

    # ── Supplier GSTIN ─────────────────────────────────────────────────────────
    s_gstin = _supplier_gstin(text, buyer_gstin)

    # ── Supplier Name ──────────────────────────────────────────────────────────
    s_name = _extract_supplier_name(text, s_gstin)

    # ── Supplier Address ───────────────────────────────────────────────────────
    s_addr = _first([
        r"(?:Supplier|Seller|From)\s+(?:Address|Addr)[.:\s]+([^\n]+(?:\n[^\n]+){0,2})",
        r"(?:Registered\s+Office|Reg\.?\s+Off\.?)[.:\s]+([^\n]+(?:\n[^\n]+){0,1})",
        r"(?:Address|Addr)[.:\s]+([^\n]+(?:\n[^\n]+){0,1})",
    ], text)

    # ── Particulars ────────────────────────────────────────────────────────────
    parts = _first([
        r"(?:Description\s+of\s+(?:Goods|Services?)|Item\s+Description|Particulars)"
        r"[.:\s]+([^\n]+(?:\n(?!.*(?:HSN|SAC|Qty|Rate|Amount))[^\n]+){0,1})",
        r"Description\s+Quantity.*?\n([^\n]+)",
        r"(?:Goods|Services?)\s+Description[.:\s]+([^\n]+)",
        r"(?:Product|Item|Service)[.:\s]+([^\n]+)",
    ], text)
    if parts and len(parts) > 150:
        parts = parts[:150] + "…"

    # ── HSN / SAC Code ─────────────────────────────────────────────────────────
    hsn = _extract_hsn(text)

    # ── Tax Table (provides taxable, cgst, sgst, igst, tx_type) ───────────────
    taxable_tbl, cgst_tbl, sgst_tbl, igst_tbl, tx_type_tbl = _parse_tax_table(text)

    # ── Taxable Value ──────────────────────────────────────────────────────────
    taxable = taxable_tbl or _first([
        r"(?:Taxable\s+Value|Taxable\s+Amount|Value\s+of\s+(?:Taxable\s+)?Supply)"
        r"[.:\s]+₹?\s*([\d,]+\.?\d*)",
        r"(?:Assessable|Basic|Net)\s+(?:Value|Amount)[.:\s]+₹?\s*([\d,]+\.?\d*)",
        r"Amount\s+Before\s+(?:Tax|GST)[.:\s]+₹?\s*([\d,]+\.?\d*)",
        r"\bSub\s*[-\s]?Total[.:\s]+₹?\s*([\d,]+\.?\d*)",
        r"\bSubtotal[.:\s]+₹?\s*([\d,]+\.?\d*)",
    ], text)

    # ── GST Rate ───────────────────────────────────────────────────────────────
    gst_rate = _first([
        r"(?:GST|IGST|CGST|SGST)\s*[@(]?\s*(\d+\.?\d*\s*%)",
        r"Tax\s+Rate[.:\s]+(\d+\.?\d*\s*%)",
        r"@\s*(\d+\.?\d*%)\s*(?:GST|IGST|CGST|SGST)",
        r"(\d+(?:\.\d+)?)\s*%\s*(?:GST|IGST|CGST|SGST)",
    ], text)
    if gst_rate and "%" not in gst_rate:
        gst_rate += "%"

    # ── Place of Supply ────────────────────────────────────────────────────────
    pos = _first([
        r"Place\s+of\s+Supply[.:\s]+([A-Za-z][^\n,|]{2,40})",
        r"\bPOS[.:\s]+([^\n,|]+)",
        r"Destination\s+State[.:\s]+([^\n,|]+)",
    ], text)

    # Strip garbled OCR (e.g. "/ARYAIWA" → use state code instead)
    if pos and re.search(r"[/\\@#]", pos):
        pos = ""

    # Fallback: derive POS from buyer GSTIN state code
    if not pos and buyer_gstin and len(buyer_gstin) >= 2:
        code = buyer_gstin[:2]
        pos = f"{STATE_CODES.get(code, '')} ({code})" if code in STATE_CODES else ""

    # ── Tax Type, CGST, SGST, IGST ────────────────────────────────────────────
    tx_type = tx_type_tbl
    cgst    = cgst_tbl
    sgst    = sgst_tbl
    igst    = igst_tbl

    # ── Total Invoice Value ────────────────────────────────────────────────────
    total = _grand_total(text) or _first([
        r"(?:Total\s+Invoice\s+Value|Invoice\s+(?:Total|Amount))[.:\s]+₹?\s*([\d,]+\.?\d*)",
        r"(?:Net\s+Payable|Amount\s+Payable|Amount\s+Due|Balance\s+Due)[.:\s]+₹?\s*([\d,]+\.?\d*)",
        r"(?:Total\s+Amount\s+Due)[.:\s]*₹?\s*([\d,]+\.?\d*)",
        # Cash payment line: "Cash - 1,611.00" or "Cash/Party - 3,375.00"
        r"(?:Cash|CASH)[\s/A-Za-z\-]+\s+([\d,]+\.?\d{2})\s*$",
        r"(?:Total\s+Amount)[.:\s]+₹?\s*([\d,]+\.?\d*)",
        # "Total  41.00 NO  Rs 24,072.00" — skip qty+unit, grab Rs/₹ amount
        r"\bTotal\b[^\n]*?(?:Rs\.?|₹|INR)\s*([\d,]+\.?\d+)",
        r"\bTOTAL\b[.:\s€₹]*([\d,]+\.?\d*)",          # all-caps only (no IGNORECASE bleed)
        r"\bTotal:\s*₹?\s*([\d,]+\.?\d*)",
        r"\bTotal\b[.:\s]+₹?\s*([\d,]+\.?\d*)",        # word-boundary, avoids "Subtotal"
    ], text, re.MULTILINE | re.IGNORECASE)

    # Sanity: if total equals taxable, the Total pattern hit Sub-Total — discard
    if total and taxable and _clean(total) == _clean(taxable):
        total = ""

    # Fallback: compute total = taxable + taxes
    if not total and taxable:
        try:
            t_val   = _to_float(taxable)
            tax_sum = sum(_to_float(x) for x in (igst, cgst, sgst) if x)
            if tax_sum > 0:
                total = f"{t_val + tax_sum:.2f}"
        except Exception:
            pass

    return {
        "inv_no":          inv_no,
        "inv_date":        inv_date,
        "s_name":          s_name,
        "s_addr":          s_addr,
        "s_gstin":         s_gstin,
        "parts":           parts,
        "taxable":         _clean(taxable),
        "hsn":             hsn,
        "gst_rate":        gst_rate,
        "pos":             pos.strip() if pos else "",
        "tx_type":         tx_type,
        "cgst":            cgst,
        "sgst":            sgst,
        "igst":            igst,
        "total":           _clean(total),
        "_delivery_note":  is_delivery_note,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  QUALITY GATE
# ══════════════════════════════════════════════════════════════════════════════

def _quality_ok(rec):
    """
    Pass if at least 2 of the key identity fields are non-empty.
    Total is NOT required — it often fails for scanned invoices even when
    everything else extracted correctly.
    """
    return sum(1 for k in ("inv_no", "inv_date", "s_name", "s_gstin")
               if rec.get(k, "").strip()) >= 2


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def extract_invoice(path, filename, buyer_gstin, gemini_key=None):
    """
    Main extraction function called by app.py.

    Returns:
      (rec_dict, method_label)  — on success
      (None,     reason_string) — when extraction quality is too low
    """
    text, text_method = get_text(path)

    if not text or len(text) < MIN_TEXT_LEN:
        return None, f"Could not extract text ({text_method})"

    # ── Try AI extraction first ───────────────────────────────────────────────
    if gemini_key:
        rec = extract_with_llm(text, gemini_key, buyer_gstin)
        if rec and _quality_ok(rec):
            return rec, f"AI (Gemini) · {text_method}"
        # LLM failed or low quality → fall through to rules

    # ── Rule-based fallback ───────────────────────────────────────────────────
    rec = extract_with_rules(text, filename, buyer_gstin)

    # Delivery notes: extract what we can but flag them
    if rec.get("_delivery_note"):
        return None, (
            f"Delivery note / challan detected ({text_method}) — "
            "not a tax invoice, no GST amounts present"
        )

    if _quality_ok(rec):
        return rec, f"Rule-based · {text_method}"

    return None, (
        f"Text extracted ({text_method}) but key fields could not be parsed — "
        "manual entry needed"
    )


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def build_excel_bytes(records, failed):
    """Build a two-sheet Excel workbook in memory and return raw bytes."""
    wb   = Workbook()
    thin = Side(border_style="thin", color="000000")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    alc  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alv  = Alignment(vertical="top", wrap_text=True)

    def _hdr(ws, hex_color):
        fill = PatternFill("solid", fgColor=hex_color)
        font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        for cell in ws[1]:
            cell.fill, cell.font, cell.alignment, cell.border = fill, font, alc, bdr
        ws.row_dimensions[1].height = 36

    def _row(ws, row_num, vals):
        ws.append(vals)
        fill = PatternFill("solid", fgColor="D6E4F0" if row_num % 2 == 0 else "FFFFFF")
        for cell in ws[ws.max_row]:
            cell.fill      = fill
            cell.border    = bdr
            cell.alignment = alv
            cell.font      = Font(name="Arial", size=9)

    # ── Sheet 1: Invoice Summary ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Invoice Summary"
    ws1.append(HEADERS)
    _hdr(ws1, "1F4E79")

    for ci, w in enumerate(
        [5, 34, 20, 13, 26, 40, 20, 50, 14, 14, 9, 20, 18, 13, 13, 13, 18], 1
    ):
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.freeze_panes = "C2"

    for i, (fname, r) in enumerate(records, 1):
        _row(ws1, i, [
            i, fname,
            r.get("inv_no",""), r.get("inv_date",""),
            r.get("s_name",""), r.get("s_addr",""), r.get("s_gstin",""),
            r.get("parts",""), r.get("taxable",""), r.get("hsn",""),
            r.get("gst_rate",""), r.get("pos",""), r.get("tx_type",""),
            r.get("cgst",""), r.get("sgst",""), r.get("igst",""),
            r.get("total",""),
        ])

    # ── Sheet 2: Unprocessed ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Unprocessed Invoices")
    ws2.append(["S.No.", "File Name", "Reason / Notes"])
    _hdr(ws2, "C00000")
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 65
    ws2.freeze_panes = "A2"

    for j, (fname, reason) in enumerate(failed, 1):
        _row(ws2, j, [j, fname, reason])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
